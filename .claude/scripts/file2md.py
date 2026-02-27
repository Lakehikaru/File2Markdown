#!/usr/bin/env python3
"""file2md.py - Convert files to Markdown via markdown.new API.

Handles files > 10MB by splitting PDF (pypdfium2), XLSX (openpyxl),
and DOCX (zipfile+xml) into chunks, converting each, then merging results.
"""

import argparse
import copy
import os
import random
import string
import sys
import tempfile
import time
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import requests

API_URL = "https://markdown.new/convert"
HOME_URL = "https://markdown.new/"
MAX_SIZE = 10 * 1024 * 1024  # 10MB


def get_api_key(cli_key: str | None) -> str | None:
    return cli_key or os.environ.get("MARKDOWN_NEW_API_KEY") or None


def create_session(api_key: str | None) -> requests.Session:
    """Create a session that mimics a real Chrome browser."""
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.6778.86 Safari/537.36",
        "Accept-Language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Connection": "keep-alive",
        "DNT": "1",
    })
    if api_key:
        s.headers["Authorization"] = f"Bearer {api_key}"

    # Visit homepage first to get cookies, like a real user
    print("Initializing session...")
    try:
        resp = s.get(HOME_URL, headers={
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
            "Upgrade-Insecure-Requests": "1",
            "Cache-Control": "max-age=0",
        }, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"Warning: failed to init session: {e}", file=sys.stderr)

    time.sleep(random.uniform(1, 3))
    return s


def random_filename(suffix: str) -> str:
    """Generate a realistic-looking filename."""
    base = ''.join(random.choices(string.ascii_lowercase + string.digits, k=random.randint(6, 12)))
    return f"document_{base}{suffix}"


def convert_single(session: requests.Session, filepath: Path) -> str:
    """Upload a single file (must be <= 10MB) and return markdown text."""
    upload_name = random_filename(filepath.suffix)
    with open(filepath, "rb") as f:
        resp = session.post(
            API_URL,
            files={"file": (upload_name, f)},
            headers={
                "Accept": "application/json, text/plain, */*",
                "Origin": "https://markdown.new",
                "Referer": "https://markdown.new/",
                "Sec-Fetch-Dest": "empty",
                "Sec-Fetch-Mode": "cors",
                "Sec-Fetch-Site": "same-origin",
                "Sec-Ch-Ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
                "Sec-Ch-Ua-Mobile": "?0",
                "Sec-Ch-Ua-Platform": '"Windows"',
            },
            timeout=180,
        )
    if resp.status_code != 200:
        print(f"API error {resp.status_code}: {resp.text}", file=sys.stderr)
        sys.exit(1)
    try:
        body = resp.json()
        data = body.get("data", body)
        return data.get("content", resp.text)
    except Exception:
        return resp.text


# --------------- PDF splitting ---------------

def split_pdf(filepath: Path, tmpdir: str) -> list[Path]:
    """Split a PDF into chunks each < 10MB using pypdfium2."""
    import pypdfium2 as pdfium

    pdf = pdfium.PdfDocument(filepath)
    total_pages = len(pdf)
    file_size = filepath.stat().st_size

    pages_per_chunk = max(1, int(total_pages * 9 * 1024 * 1024 / file_size))

    chunks: list[Path] = []
    start = 0
    while start < total_pages:
        end = min(start + pages_per_chunk, total_pages)
        chunk_path = Path(tmpdir) / f"chunk_{len(chunks):04d}.pdf"

        new_pdf = pdfium.PdfDocument.new()
        new_pdf.import_pages(pdf, list(range(start, end)))
        new_pdf.save(str(chunk_path))
        new_pdf.close()

        while chunk_path.stat().st_size > MAX_SIZE and end - start > 1:
            end = start + (end - start) // 2
            new_pdf = pdfium.PdfDocument.new()
            new_pdf.import_pages(pdf, list(range(start, end)))
            new_pdf.save(str(chunk_path))
            new_pdf.close()

        if chunk_path.stat().st_size > MAX_SIZE:
            print(f"Warning: single page {start} exceeds 10MB, uploading anyway.", file=sys.stderr)

        chunks.append(chunk_path)
        start = end

    pdf.close()
    return chunks


# --------------- XLSX splitting ---------------

def split_xlsx(filepath: Path, tmpdir: str) -> list[Path]:
    """Split an XLSX by sheet, each sheet as a separate file."""
    from openpyxl import load_workbook, Workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    chunks: list[Path] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        for row in ws.iter_rows(values_only=True):
            new_ws.append(list(row))

        chunk_path = Path(tmpdir) / f"{sheet_name}.xlsx"
        new_wb.save(str(chunk_path))

        if chunk_path.stat().st_size > MAX_SIZE:
            print(f"Warning: sheet '{sheet_name}' exceeds 10MB. API may reject it.", file=sys.stderr)

        chunks.append(chunk_path)

    wb.close()
    return chunks


# --------------- DOCX splitting ---------------

DOCX_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

def split_docx(filepath: Path, tmpdir: str) -> list[Path]:
    """Split a DOCX by paragraphs into chunks each < 10MB using zipfile+xml."""
    # Read original docx as zip
    with zipfile.ZipFile(filepath, "r") as zin:
        names = zin.namelist()
        file_contents = {n: zin.read(n) for n in names}

    # Parse document.xml to get paragraphs
    doc_xml = file_contents["word/document.xml"]
    ET.register_namespace("w", DOCX_NS)
    # Preserve all namespaces from original
    ns_map = {}
    for event, elem in ET.iterparse(filepath.open("rb") if False else __import__("io").BytesIO(doc_xml), events=["start-ns"]):
        prefix, uri = elem
        if prefix:
            ET.register_namespace(prefix, uri)
        ns_map[prefix] = uri

    tree = ET.parse(__import__("io").BytesIO(doc_xml))
    root = tree.getroot()
    body = root.find(f"{{{DOCX_NS}}}body")
    if body is None:
        # Try without namespace
        body = root.find(".//body") or root[0]

    children = list(body)
    # sectPr (section properties) is usually the last element, keep it in every chunk
    sect_pr = None
    if children and children[-1].tag.endswith("}sectPr"):
        sect_pr = children.pop()

    if not children:
        return [filepath]

    # Estimate paragraphs per chunk based on file size ratio
    file_size = filepath.stat().st_size
    total_paras = len(children)
    paras_per_chunk = max(1, int(total_paras * 8 * 1024 * 1024 / file_size))

    chunks: list[Path] = []
    start = 0
    while start < total_paras:
        end = min(start + paras_per_chunk, total_paras)

        # Build new document.xml with subset of paragraphs
        new_root = copy.deepcopy(root)
        new_body = new_root.find(f"{{{DOCX_NS}}}body")
        if new_body is None:
            new_body = new_root[0]

        # Remove all children from body
        for child in list(new_body):
            new_body.remove(child)

        # Add selected paragraphs
        for child in children[start:end]:
            new_body.append(copy.deepcopy(child))

        # Add section properties
        if sect_pr is not None:
            new_body.append(copy.deepcopy(sect_pr))

        # Write new docx
        chunk_path = Path(tmpdir) / f"chunk_{len(chunks):04d}.docx"
        new_xml = ET.tostring(new_root, xml_declaration=True, encoding="UTF-8")

        with zipfile.ZipFile(chunk_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                if name == "word/document.xml":
                    zout.writestr(name, new_xml)
                else:
                    zout.writestr(name, file_contents[name])

        # Shrink if too large
        while chunk_path.stat().st_size > MAX_SIZE and end - start > 1:
            end = start + (end - start) // 2
            new_root2 = copy.deepcopy(root)
            new_body2 = new_root2.find(f"{{{DOCX_NS}}}body")
            if new_body2 is None:
                new_body2 = new_root2[0]
            for child in list(new_body2):
                new_body2.remove(child)
            for child in children[start:end]:
                new_body2.append(copy.deepcopy(child))
            if sect_pr is not None:
                new_body2.append(copy.deepcopy(sect_pr))
            new_xml2 = ET.tostring(new_root2, xml_declaration=True, encoding="UTF-8")
            with zipfile.ZipFile(chunk_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for name in names:
                    if name == "word/document.xml":
                        zout.writestr(name, new_xml2)
                    else:
                        zout.writestr(name, file_contents[name])

        if chunk_path.stat().st_size > MAX_SIZE:
            print(f"Warning: chunk starting at paragraph {start} exceeds 10MB.", file=sys.stderr)

        chunks.append(chunk_path)
        start = end

    return chunks


# --------------- Main ---------------

def convert_file(session: requests.Session, input_path: Path) -> str:
    """Convert a file to markdown, splitting if necessary."""
    file_size = input_path.stat().st_size
    suffix = input_path.suffix.lower()

    if file_size <= MAX_SIZE:
        print(f"File size {file_size / 1024 / 1024:.1f}MB, uploading directly...")
        return convert_single(session, input_path)

    print(f"File size {file_size / 1024 / 1024:.1f}MB exceeds 10MB, splitting...")

    with tempfile.TemporaryDirectory() as tmpdir:
        if suffix == ".pdf":
            chunks = split_pdf(input_path, tmpdir)
        elif suffix == ".xlsx":
            chunks = split_xlsx(input_path, tmpdir)
        elif suffix in (".docx", ".doc"):
            chunks = split_docx(input_path, tmpdir)
        else:
            print(f"Error: auto-split not supported for {suffix} files > 10MB.", file=sys.stderr)
            print("Please manually split the file into parts < 10MB.", file=sys.stderr)
            sys.exit(1)

        print(f"Split into {len(chunks)} chunks.")
        parts: list[str] = []
        for i, chunk in enumerate(chunks):
            if i > 0:
                delay = random.uniform(8, 20)
                print(f"Waiting {delay:.0f}s...")
                time.sleep(delay)
            print(f"Converting chunk {i + 1}/{len(chunks)}: {chunk.name} ({chunk.stat().st_size / 1024 / 1024:.1f}MB)...")
            md = convert_single(session, chunk)
            parts.append(md)

    return "\n\n---\n\n".join(parts)


def main():
    parser = argparse.ArgumentParser(description="Convert files to Markdown via markdown.new API")
    parser.add_argument("input_file", nargs="+", help="Input file path(s)")
    parser.add_argument("--output", "-o", help="Output .md file path (only for single file)")
    parser.add_argument("--api-key", "-k", help="markdown.new API key")
    args = parser.parse_args()

    api_key = get_api_key(args.api_key)
    session = create_session(api_key)

    for idx, fpath in enumerate(args.input_file):
        if idx > 0:
            delay = random.uniform(15, 30)
            print(f"\nWaiting {delay:.0f}s before next file...")
            time.sleep(delay)

        input_path = Path(fpath)
        if not input_path.exists():
            print(f"Error: file not found: {input_path}", file=sys.stderr)
            continue

        output_path = Path(args.output) if (args.output and len(args.input_file) == 1) else input_path.with_suffix(".md")

        print(f"\n[{idx + 1}/{len(args.input_file)}] Converting: {input_path.name}")
        markdown = convert_file(session, input_path)

        output_path.write_text(markdown, encoding="utf-8")
        print(f"Done! Output: {output_path}")


if __name__ == "__main__":
    main()
